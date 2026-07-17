from fastapi import APIRouter, Depends, HTTPException, status, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.orm import Session
from app.api.deps import require_roles, get_current_user
from app.core.database import get_db
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import tempfile
from app.core.time import utc_now
from app.models import BloqueHorasExtra, CorreccionManual, Departamento, DetallePlantillaTurno, Empleado, PlantillaTurno, RegistroAsistencia, RegistroAusencia, RolNombre, SalidaTemporal, TurnoHorario, UsuarioSistema, Rol, SupervisorDepartamento, Visita, EstadoVisita, EstadoRegistro, TipoCorreccion
from app.schemas import AusenciaCreate, AusenciaOut, CorreccionManualCreate, CorreccionManualOut, EmpleadoCreate, EmpleadoOut, EmpleadoUpdate, TurnoCreate, TurnoOut, TurnoUpdate, ImportarEmpleadosResponse


router = APIRouter(prefix="/admin", tags=["admin"])
ADMIN_ACCESS = Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))


@router.post("/empleados", response_model=EmpleadoOut, dependencies=[ADMIN_ACCESS])
def crear_empleado(payload: EmpleadoCreate, db: Session = Depends(get_db)):
    empleado = Empleado(**payload.model_dump())
    db.add(empleado)
    db.commit()
    db.refresh(empleado)
    return empleado


@router.post("/empleados/importar-excel", response_model=ImportarEmpleadosResponse, dependencies=[ADMIN_ACCESS])
def importar_empleados_excel(
    archivo: UploadFile,
    db: Session = Depends(get_db)
):
    """Importa colaboradores desde un archivo Excel.
    
    Formato del Excel:
    - Columna A: numero_empleado (obligatorio)
    - Columna B: nombre_completo (obligatorio)
    - Columna C: departamento_id (obligatorio)
    - Columna D: puesto (obligatorio)
    """
    if not archivo.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser un Excel (.xlsx o .xls)")
    
    try:
        # Leer el archivo Excel
        contenido = archivo.file.read()
        print(f"Archivo leído: {len(contenido)} bytes")
        
        # Usar BytesIO para leer el contenido
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(contenido), read_only=True)
        hoja = wb.active
        print(f"Hoja activa: {hoja.title}")
        
        exitosos = 0
        errores = []
        total_procesados = 0
        
        # Obtener departamentos existentes para validar
        departamentos = db.scalars(select(Departamento)).all()
        deptos_por_id = {d.id: d.nombre for d in departamentos}
        deptos_por_nombre = {d.nombre.lower(): d.id for d in departamentos}
        
        # Empezar desde la fila 2 (asumiendo fila 1 como encabezados)
        for fila_idx, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            total_procesados += 1
            
            # Obtener valores de las columnas
            numero_empleado = str(fila[0]).strip() if fila[0] else None
            nombre_completo = str(fila[1]).strip() if fila[1] else None
            departamento = str(fila[2]).strip() if fila[2] else None
            puesto = str(fila[3]).strip() if fila[3] else None
            
            # Validaciones
            if not numero_empleado or not nombre_completo or not departamento or not puesto:
                errores.append({
                    "fila": fila_idx,
                    "mensaje": "Faltan datos obligatorios (numero_empleado, nombre_completo, departamento, puesto)"
                })
                continue
            
            # Verificar si el número de empleado ya existe
            existente = db.scalar(
                select(Empleado).where(Empleado.numero_empleado == numero_empleado)
            )
            if existente:
                errores.append({
                    "fila": fila_idx,
                    "mensaje": f"El número de empleado {numero_empleado} ya existe"
                })
                continue
            
            # Obtener departamento_id
            departamento_id = None
            try:
                departamento_id = int(departamento)
                if departamento_id not in deptos_por_id:
                    errores.append({
                        "fila": fila_idx,
                        "mensaje": f"Departamento ID {departamento_id} no existe"
                    })
                    continue
            except ValueError:
                # Intentar buscar por nombre
                depto_id = deptos_por_nombre.get(departamento.lower())
                if not depto_id:
                    errores.append({
                        "fila": fila_idx,
                        "mensaje": f"Departamento '{departamento}' no encontrado"
                    })
                    continue
                departamento_id = depto_id
            
            # Crear empleado
            try:
                empleado = Empleado(
                    numero_empleado=numero_empleado,
                    nombre_completo=nombre_completo,
                    departamento_id=departamento_id,
                    puesto=puesto,
                    activo=True
                )
                db.add(empleado)
                db.commit()
                exitosos += 1
            except Exception as e:
                db.rollback()
                errores.append({
                    "fila": fila_idx,
                    "mensaje": f"Error al crear empleado: {str(e)}"
                })
        
        return ImportarEmpleadosResponse(
            exitosos=exitosos,
            errores=errores,
            total_procesados=total_procesados
        )
        
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al procesar el archivo: {str(e)}")


@router.get("/empleados", response_model=list[EmpleadoOut], dependencies=[ADMIN_ACCESS])
def listar_empleados(db: Session = Depends(get_db)):
    try:
        return db.scalars(select(Empleado).order_by(Empleado.nombre_completo)).all()
    except Exception as e:
        print(f"Error en listar_empleados: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al obtener colaboradores")


@router.get("/empleados/{id}", response_model=EmpleadoOut, dependencies=[ADMIN_ACCESS])
def obtener_empleado(id: int, db: Session = Depends(get_db)):
    empleado = db.get(Empleado, id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    return empleado


@router.put("/empleados/{id}", response_model=EmpleadoOut, dependencies=[ADMIN_ACCESS])
def actualizar_empleado(id: int, payload: EmpleadoUpdate, db: Session = Depends(get_db)):
    empleado = db.get(Empleado, id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(empleado, key, value)
    db.commit()
    db.refresh(empleado)
    return empleado


@router.delete("/empleados/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def eliminar_empleado(id: int, db: Session = Depends(get_db)):
    empleado = db.get(Empleado, id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    db.delete(empleado)
    db.commit()
    return {"ok": True}


@router.put("/empleados/{id}/activo", dependencies=[ADMIN_ACCESS])
def activar_desactivar_empleado(id: int, activo: bool, db: Session = Depends(get_db)):
    empleado = db.get(Empleado, id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    empleado.activo = activo
    db.commit()
    db.refresh(empleado)
    return {"ok": True, "activo": empleado.activo}


@router.post("/turnos", response_model=TurnoOut, dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def crear_turno(payload: TurnoCreate, db: Session = Depends(get_db)):
    turno = TurnoHorario(**payload.model_dump())
    db.add(turno)
    db.commit()
    db.refresh(turno)
    return turno


@router.get("/turnos", response_model=list[TurnoOut], dependencies=[ADMIN_ACCESS])
def listar_turnos(db: Session = Depends(get_db)):
    try:
        return db.scalars(select(TurnoHorario).order_by(TurnoHorario.empleado_id, TurnoHorario.dia_semana)).all()
    except Exception as e:
        print(f"Error en listar_turnos: {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al obtener turnos")


@router.get("/turnos-raw", dependencies=[ADMIN_ACCESS])
def listar_turnos_raw(db: Session = Depends(get_db)):
    turnos = db.scalars(select(TurnoHorario).order_by(TurnoHorario.empleado_id, TurnoHorario.dia_semana)).all()
    return [
        {
            "id": t.id,
            "empleado_id": t.empleado_id,
            "dia_semana": t.dia_semana,
            "hora_entrada_oficial": str(t.hora_entrada_oficial) if t.hora_entrada_oficial else None,
            "hora_salida_oficial": str(t.hora_salida_oficial) if t.hora_salida_oficial else None,
            "tolerancia_minutos": t.tolerancia_minutos,
            "es_descanso": t.es_descanso
        }
        for t in turnos
    ]


@router.put("/turnos/{id}", response_model=TurnoOut, dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def actualizar_turno(id: int, payload: TurnoUpdate, db: Session = Depends(get_db)):
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Actualizando turno {id} con payload: {payload.model_dump(exclude_unset=True)}")
    
    turno = db.get(TurnoHorario, id)
    if not turno:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Turno no encontrado")
    
    for key, value in payload.model_dump(exclude_unset=True).items():
        logger.info(f"Actualizando campo {key} con valor {value}")
        setattr(turno, key, value)
    
    db.commit()
    db.refresh(turno)
    logger.info(f"Turno actualizado: entrada={turno.hora_entrada_oficial}, salida={turno.hora_salida_oficial}")
    return turno


@router.delete("/turnos/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def eliminar_turno(id: int, db: Session = Depends(get_db)):
    turno = db.get(TurnoHorario, id)
    if not turno:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Turno no encontrado")
    db.delete(turno)
    db.commit()
    return {"ok": True}


# Plantillas de turnos
@router.post("/plantillas-turnos", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def crear_plantilla(
    nombre: str,
    descripcion: str | None = None,
    es_rotativa: bool = False,
    ciclo_rotacion_semanas: int = 2,
    fecha_inicio_ciclo: str | None = None,
    plantilla_semana_par_id: int | None = None,
    plantilla_semana_impar_id: int | None = None,
    plantilla_semana_3_id: int | None = None,
    db: Session = Depends(get_db)
):
    from datetime import date
    if fecha_inicio_ciclo:
        fecha_inicio = date.fromisoformat(fecha_inicio_ciclo)
    else:
        fecha_inicio = date.today()
    plantilla = PlantillaTurno(
        nombre=nombre,
        descripcion=descripcion,
        es_rotativa=es_rotativa,
        ciclo_rotacion_semanas=ciclo_rotacion_semanas,
        fecha_inicio_ciclo=fecha_inicio,
        plantilla_semana_par_id=plantilla_semana_par_id,
        plantilla_semana_impar_id=plantilla_semana_impar_id,
        plantilla_semana_3_id=plantilla_semana_3_id
    )
    db.add(plantilla)
    db.commit()
    db.refresh(plantilla)
    return plantilla


@router.get("/plantillas-turnos", dependencies=[ADMIN_ACCESS])
def listar_plantillas(db: Session = Depends(get_db)):
    return db.scalars(select(PlantillaTurno).order_by(PlantillaTurno.nombre)).all()


@router.delete("/plantillas-turnos/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def eliminar_plantilla(id: int, db: Session = Depends(get_db)):
    plantilla = db.get(PlantillaTurno, id)
    if not plantilla:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla no encontrada")
    # Eliminar detalles primero
    db.scalars(select(DetallePlantillaTurno).where(DetallePlantillaTurno.plantilla_id == id)).all()
    for detalle in db.scalars(select(DetallePlantillaTurno).where(DetallePlantillaTurno.plantilla_id == id)):
        db.delete(detalle)
    db.delete(plantilla)
    db.commit()
    return {"ok": True}


@router.put("/plantillas-turnos/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def actualizar_plantilla(
    id: int,
    nombre: str,
    descripcion: str | None = None,
    es_rotativa: bool | None = None,
    ciclo_rotacion_semanas: int | None = None,
    fecha_inicio_ciclo: str | None = None,
    plantilla_semana_par_id: int | None = None,
    plantilla_semana_impar_id: int | None = None,
    plantilla_semana_3_id: int | None = None,
    db: Session = Depends(get_db)
):
    plantilla = db.get(PlantillaTurno, id)
    if not plantilla:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla no encontrada")
    plantilla.nombre = nombre
    if descripcion is not None:
        plantilla.descripcion = descripcion
    if es_rotativa is not None:
        plantilla.es_rotativa = es_rotativa
    if ciclo_rotacion_semanas is not None:
        plantilla.ciclo_rotacion_semanas = ciclo_rotacion_semanas
    if fecha_inicio_ciclo is not None:
        from datetime import date
        plantilla.fecha_inicio_ciclo = date.fromisoformat(fecha_inicio_ciclo)
    if plantilla_semana_par_id is not None:
        plantilla.plantilla_semana_par_id = plantilla_semana_par_id
    if plantilla_semana_impar_id is not None:
        plantilla.plantilla_semana_impar_id = plantilla_semana_impar_id
    if plantilla_semana_3_id is not None:
        plantilla.plantilla_semana_3_id = plantilla_semana_3_id
    db.commit()
    db.refresh(plantilla)
    return plantilla


@router.post("/plantillas-turnos/{plantilla_id}/detalles", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def agregar_detalle_plantilla(plantilla_id: int, data: dict, db: Session = Depends(get_db)):
    from datetime import time
    plantilla = db.get(PlantillaTurno, plantilla_id)
    if not plantilla:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla no encontrada")

    dia_semana = data.get('dia_semana')
    hora_entrada = data.get('hora_entrada')
    hora_salida = data.get('hora_salida')
    tolerancia = data.get('tolerancia', 15)
    es_descanso = data.get('es_descanso', False)
    es_por_asistencia = data.get('es_por_asistencia', False)

    detalle = DetallePlantillaTurno(
        plantilla_id=plantilla_id,
        dia_semana=dia_semana,
        hora_entrada_oficial=time.fromisoformat(hora_entrada) if hora_entrada and not es_por_asistencia else None,
        hora_salida_oficial=time.fromisoformat(hora_salida) if hora_salida and not es_por_asistencia else None,
        tolerancia_minutos=tolerancia,
        es_descanso=es_descanso,
        es_por_asistencia=es_por_asistencia
    )
    db.add(detalle)
    db.commit()
    db.refresh(detalle)
    return detalle


@router.get("/empleados/{empleado_id}/plantilla-efectiva", dependencies=[ADMIN_ACCESS])
def obtener_plantilla_efectiva(
    empleado_id: int,
    fecha: str | None = None,
    db: Session = Depends(get_db)
):
    """Obtiene la plantilla efectiva para un empleado en una fecha específica, considerando rotación."""
    from datetime import datetime
    from app.services import get_empleado_turno
    
    empleado = db.get(Empleado, empleado_id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empleado no encontrado")
    
    if not empleado.plantilla_turno_id:
        return {"plantilla_efectiva": None, "es_rotativa": False}
    
    plantilla = db.get(PlantillaTurno, empleado.plantilla_turno_id)
    if not plantilla:
        return {"plantilla_efectiva": None, "es_rotativa": False}
    
    # Si la plantilla es rotativa, determinar cuál usar según fecha
    plantilla_efectiva = plantilla
    if fecha:
        fecha_obj = datetime.fromisoformat(fecha).date()
    else:
        fecha_obj = utc_now().date()

    semana_ciclo = None
    if plantilla.es_rotativa:
        ciclo = plantilla.ciclo_rotacion_semanas
        # Fecha de inicio del ciclo: por empleado o por plantilla
        fecha_inicio = empleado.fecha_inicio_ciclo or plantilla.fecha_inicio_ciclo
        dias_desde_inicio = (fecha_obj - fecha_inicio).days
        semanas_transcurridas = dias_desde_inicio // 7
        semana_ciclo = (semanas_transcurridas % ciclo) + 1
        if semana_ciclo == 3 and ciclo == 3:
            plantilla_efectiva = plantilla.plantilla_semana_3
        elif semana_ciclo == 2:
            plantilla_efectiva = plantilla.plantilla_semana_par
        else:
            plantilla_efectiva = plantilla.plantilla_semana_impar

    return {
        "plantilla_efectiva": {
            "id": plantilla_efectiva.id,
            "nombre": plantilla_efectiva.nombre,
            "descripcion": plantilla_efectiva.descripcion
        } if plantilla_efectiva else None,
        "es_rotativa": plantilla.es_rotativa,
        "ciclo_rotacion_semanas": plantilla.ciclo_rotacion_semanas,
        "semana_ciclo": semana_ciclo,
        "fecha_inicio_ciclo": plantilla.fecha_inicio_ciclo.isoformat(),
        "fecha_consulta": fecha_obj.isoformat(),
        "plantilla_base": {
            "id": plantilla.id,
            "nombre": plantilla.nombre
        }
    }


@router.get("/plantillas-turnos/{plantilla_id}/detalles", dependencies=[ADMIN_ACCESS])
def listar_detalles_plantilla(plantilla_id: int, db: Session = Depends(get_db)):
    detalles = db.scalars(select(DetallePlantillaTurno).where(DetallePlantillaTurno.plantilla_id == plantilla_id).order_by(DetallePlantillaTurno.dia_semana)).all()
    return [
        {
            "id": d.id,
            "plantilla_id": d.plantilla_id,
            "dia_semana": d.dia_semana,
            "hora_entrada_oficial": d.hora_entrada_oficial.isoformat() if d.hora_entrada_oficial else None,
            "hora_salida_oficial": d.hora_salida_oficial.isoformat() if d.hora_salida_oficial else None,
            "tolerancia_minutos": d.tolerancia_minutos,
            "es_descanso": d.es_descanso,
            "es_por_asistencia": d.es_por_asistencia
        }
        for d in detalles
    ]


@router.get("/plantillas-turnos/{plantilla_id}/info", dependencies=[ADMIN_ACCESS])
def info_plantilla(plantilla_id: int, db: Session = Depends(get_db)):
    plantilla = db.get(PlantillaTurno, plantilla_id)
    if not plantilla:
        return {"error": "Plantilla no encontrada"}
    detalles = db.scalars(select(DetallePlantillaTurno).where(DetallePlantillaTurno.plantilla_id == plantilla_id).order_by(DetallePlantillaTurno.dia_semana)).all()
    return {
        "id": plantilla.id,
        "nombre": plantilla.nombre,
        "descripcion": plantilla.descripcion,
        "num_detalles": len(detalles),
        "detalles": [
            {
                "dia_semana": d.dia_semana,
                "hora_entrada": str(d.hora_entrada_oficial) if d.hora_entrada_oficial else None,
                "hora_salida": str(d.hora_salida_oficial) if d.hora_salida_oficial else None,
                "tolerancia": d.tolerancia_minutos,
                "es_descanso": d.es_descanso
            }
            for d in detalles
        ]
    }


@router.delete("/plantillas-turnos/{plantilla_id}/detalles/{detalle_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def eliminar_detalle_plantilla(plantilla_id: int, detalle_id: int, db: Session = Depends(get_db)):
    detalle = db.get(DetallePlantillaTurno, detalle_id)
    if not detalle or detalle.plantilla_id != plantilla_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Detalle no encontrado")
    db.delete(detalle)
    db.commit()
    return {"ok": True}


@router.get("/usuarios-sistema", dependencies=[ADMIN_ACCESS])
def listar_usuarios_sistema(db: Session = Depends(get_db)):
    usuarios = db.scalars(select(UsuarioSistema).order_by(UsuarioSistema.id)).all()
    return [
        {
            "id": u.id,
            "username": u.username,
            "rol": u.rol.nombre,
            "rol_id": u.rol_id,
            "empleado_id": u.empleado_id,
            "empleado_nombre": u.empleado.nombre_completo if u.empleado else None,
            "activo": u.activo
        }
        for u in usuarios
    ]


@router.post("/usuarios-sistema", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def crear_usuario_sistema(username: str, password: str, rol_id: int, empleado_id: int | None = None, db: Session = Depends(get_db), current_user: UsuarioSistema = Depends(get_current_user)):
    from app.core.security import hash_password
    if db.scalar(select(UsuarioSistema).where(UsuarioSistema.username == username)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Username ya existe")
    rol = db.get(Rol, rol_id)
    if not rol:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rol no encontrado")

    # Validar restricciones de jerarquía: superusuario > admin > rrhh > supervisor
    if rol.nombre == RolNombre.SUPERUSUARIO and current_user.rol.nombre != RolNombre.SUPERUSUARIO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario puede crear superusuarios")
    if rol.nombre == RolNombre.ADMINISTRADOR and current_user.rol.nombre != RolNombre.SUPERUSUARIO:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario puede crear administradores")
    if rol.nombre == RolNombre.RRHH and current_user.rol.nombre not in [RolNombre.SUPERUSUARIO, RolNombre.ADMINISTRADOR]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario o administrador pueden crear usuarios RRHH")
    if rol.nombre == RolNombre.SUPERVISOR and current_user.rol.nombre not in [RolNombre.SUPERUSUARIO, RolNombre.ADMINISTRADOR, RolNombre.RRHH]:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo RRHH, administrador o superusuario pueden crear supervisores")

    if empleado_id:
        empleado = db.get(Empleado, empleado_id)
        if not empleado:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    usuario = UsuarioSistema(
        username=username,
        password_hash=hash_password(password),
        rol_id=rol_id,
        empleado_id=empleado_id,
        activo=True
    )
    db.add(usuario)
    db.commit()
    db.refresh(usuario)
    return {"ok": True, "id": usuario.id}


@router.put("/usuarios-sistema/{usuario_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def actualizar_usuario_sistema(usuario_id: int, username: str | None = None, rol_id: int | None = None, empleado_id: int | None = None, db: Session = Depends(get_db), current_user: UsuarioSistema = Depends(get_current_user)):
    usuario = db.get(UsuarioSistema, usuario_id)
    if not usuario:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    if username and username != usuario.username:
        if db.scalar(select(UsuarioSistema).where(UsuarioSistema.username == username)):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Username ya existe")
        usuario.username = username
    if rol_id:
        rol = db.get(Rol, rol_id)
        if not rol:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Rol no encontrado")

        # Validar restricciones de jerarquía: superusuario > admin > rrhh > supervisor
        if rol.nombre == RolNombre.SUPERUSUARIO and current_user.rol.nombre != RolNombre.SUPERUSUARIO:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario puede asignar rol de superusuario")
        if rol.nombre == RolNombre.ADMINISTRADOR and current_user.rol.nombre != RolNombre.SUPERUSUARIO:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario puede asignar rol de administrador")
        if rol.nombre == RolNombre.RRHH and current_user.rol.nombre not in [RolNombre.SUPERUSUARIO, RolNombre.ADMINISTRADOR]:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo el superusuario o administrador pueden asignar rol RRHH")
        if rol.nombre == RolNombre.SUPERVISOR and current_user.rol.nombre not in [RolNombre.SUPERUSUARIO, RolNombre.ADMINISTRADOR, RolNombre.RRHH]:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Solo RRHH, administrador o superusuario pueden asignar rol de supervisor")

        usuario.rol_id = rol_id
    if empleado_id is not None:
        if empleado_id:
            empleado = db.get(Empleado, empleado_id)
            if not empleado:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
        usuario.empleado_id = empleado_id
    db.commit()
    db.refresh(usuario)
    return {"ok": True}


@router.delete("/usuarios-sistema/{usuario_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def eliminar_usuario_sistema(usuario_id: int, db: Session = Depends(get_db)):
    usuario = db.get(UsuarioSistema, usuario_id)
    if not usuario:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    db.delete(usuario)
    db.commit()
    return {"ok": True}


@router.put("/usuarios-sistema/{usuario_id}/activo", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def activar_desactivar_usuario(usuario_id: int, activo: bool, db: Session = Depends(get_db)):
    usuario = db.get(UsuarioSistema, usuario_id)
    if not usuario:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    usuario.activo = activo
    db.commit()
    db.refresh(usuario)
    return {"ok": True, "activo": usuario.activo}


@router.get("/supervisores-departamentos", dependencies=[ADMIN_ACCESS])
def listar_supervisores_departamentos(db: Session = Depends(get_db)):
    relaciones = db.scalars(select(SupervisorDepartamento).order_by(SupervisorDepartamento.usuario_id, SupervisorDepartamento.departamento_id)).all()
    return [
        {
            "id": r.id,
            "usuario_id": r.usuario_id,
            "departamento_id": r.departamento_id,
            "usuario_nombre": r.usuario_sistema.username if r.usuario_sistema else None,
            "departamento_nombre": r.departamento.nombre if r.departamento else None
        }
        for r in relaciones
    ]


@router.post("/supervisores-departamentos", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def asignar_supervisor_departamento(usuario_id: int, departamento_id: int, db: Session = Depends(get_db)):
    usuario = db.get(UsuarioSistema, usuario_id)
    if not usuario:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    departamento = db.get(Departamento, departamento_id)
    if not departamento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Departamento no encontrado")

    # Verificar si ya existe la relación
    existente = db.scalar(
        select(SupervisorDepartamento).where(
            SupervisorDepartamento.usuario_id == usuario_id,
            SupervisorDepartamento.departamento_id == departamento_id
        )
    )
    if existente:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Relación ya existe")

    relacion = SupervisorDepartamento(usuario_id=usuario_id, departamento_id=departamento_id)
    db.add(relacion)
    db.commit()
    db.refresh(relacion)
    return {"ok": True}


@router.delete("/supervisores-departamentos/{relacion_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def eliminar_supervisor_departamento(relacion_id: int, db: Session = Depends(get_db)):
    relacion = db.get(SupervisorDepartamento, relacion_id)
    if not relacion:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Relación no encontrada")
    db.delete(relacion)
    db.commit()
    return {"ok": True}


@router.put("/plantillas-turnos/{plantilla_id}/detalles/{detalle_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def actualizar_detalle_plantilla(plantilla_id: int, detalle_id: int, data: dict, db: Session = Depends(get_db)):
    from datetime import time
    detalle = db.get(DetallePlantillaTurno, detalle_id)
    if not detalle or detalle.plantilla_id != plantilla_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Detalle no encontrado")

    es_por_asistencia = data.get('es_por_asistencia', False)
    es_descanso = data.get('es_descanso', False)
    hora_entrada = data.get('hora_entrada')
    hora_salida = data.get('hora_salida')
    tolerancia = data.get('tolerancia', 15)

    if es_por_asistencia:
        detalle.hora_entrada_oficial = None
        detalle.hora_salida_oficial = None
    else:
        if hora_entrada:
            detalle.hora_entrada_oficial = time.fromisoformat(hora_entrada)
        if hora_salida:
            detalle.hora_salida_oficial = time.fromisoformat(hora_salida)

    detalle.tolerancia_minutos = tolerancia
    detalle.es_descanso = es_descanso
    detalle.es_por_asistencia = es_por_asistencia
    db.commit()
    db.refresh(detalle)
    return detalle


@router.put("/empleados/{empleado_id}/plantilla-turno/{plantilla_id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def asignar_plantilla_empleado(
    empleado_id: int,
    plantilla_id: int,
    fecha_inicio_ciclo: str | None = None,
    db: Session = Depends(get_db)
):
    empleado = db.get(Empleado, empleado_id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    plantilla = db.get(PlantillaTurno, plantilla_id)
    if not plantilla:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla no encontrada")

    # Eliminar turnos individuales existentes (ya que usará plantilla)
    for turno in db.scalars(select(TurnoHorario).where(TurnoHorario.empleado_id == empleado_id)):
        db.delete(turno)

    # Asignar plantilla y, si es rotativa, la fecha de inicio del ciclo para este colaborador
    empleado.plantilla_turno_id = plantilla_id
    if fecha_inicio_ciclo:
        from datetime import date
        empleado.fecha_inicio_ciclo = date.fromisoformat(fecha_inicio_ciclo)
    db.commit()
    return {"ok": True}


@router.post("/empleados/{empleado_id}/romper-plantilla", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def romper_plantilla_empleado(empleado_id: int, db: Session = Depends(get_db)):
    """Rompe la referencia a la plantilla y crea turnos individuales basados en la plantilla actual."""
    empleado = db.get(Empleado, empleado_id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    
    if not empleado.plantilla_turno_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El empleado no tiene plantilla asignada")
    
    plantilla = db.get(PlantillaTurno, empleado.plantilla_turno_id)
    if not plantilla:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla no encontrada")
    
    # Si la plantilla es rotativa, usar la plantilla efectiva de la semana actual
    plantilla_base = plantilla
    if plantilla.es_rotativa:
        fecha_actual = utc_now().date()
        ciclo = plantilla.ciclo_rotacion_semanas
        # Fecha de inicio del ciclo: por empleado o por plantilla
        fecha_inicio = empleado.fecha_inicio_ciclo or plantilla.fecha_inicio_ciclo
        dias_desde_inicio = (fecha_actual - fecha_inicio).days
        semanas_transcurridas = dias_desde_inicio // 7
        semana_ciclo = (semanas_transcurridas % ciclo) + 1
        if semana_ciclo == 3 and ciclo == 3:
            plantilla = plantilla.plantilla_semana_3
        elif semana_ciclo == 2:
            plantilla = plantilla.plantilla_semana_par
        else:
            plantilla = plantilla.plantilla_semana_impar

        if not plantilla:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se encontró la plantilla efectiva para la semana actual")
    
    # Crear o actualizar turnos individuales basados en la plantilla efectiva
    detalles = db.scalars(select(DetallePlantillaTurno).where(DetallePlantillaTurno.plantilla_id == plantilla.id)).all()
    for detalle in detalles:
        if not detalle.es_descanso and detalle.hora_entrada_oficial and detalle.hora_salida_oficial:
            # Verificar si ya existe un turno para este empleado en este día
            turno_existente = db.scalar(
                select(TurnoHorario).where(
                    TurnoHorario.empleado_id == empleado_id,
                    TurnoHorario.dia_semana == detalle.dia_semana
                )
            )
            if turno_existente:
                # Actualizar turno existente
                turno_existente.hora_entrada_oficial = detalle.hora_entrada_oficial
                turno_existente.hora_salida_oficial = detalle.hora_salida_oficial
                turno_existente.tolerancia_minutos = detalle.tolerancia_minutos
                turno_existente.es_descanso = False
            else:
                # Crear nuevo turno
                nuevo_turno = TurnoHorario(
                    empleado_id=empleado_id,
                    dia_semana=detalle.dia_semana,
                    hora_entrada_oficial=detalle.hora_entrada_oficial,
                    hora_salida_oficial=detalle.hora_salida_oficial,
                    tolerancia_minutos=detalle.tolerancia_minutos,
                    es_descanso=False
                )
                db.add(nuevo_turno)
    
    # Romper la referencia a la plantilla
    empleado.plantilla_turno_id = None
    db.commit()
    return {"ok": True}


@router.put("/asistencias/{asistencia_id}/autorizar-horas-extra", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def autorizar_horas_extra(asistencia_id: int, db: Session = Depends(get_db)):
    """Autoriza las horas extra de un registro de asistencia (solo RRHH)"""
    asistencia = db.get(RegistroAsistencia, asistencia_id)
    if not asistencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro de asistencia no encontrado")
    
    if asistencia.minutos_extra_calculados == 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay horas extra calculadas para autorizar")
    
    asistencia.autorizacion_horas_extra_rrhh = True
    db.commit()
    return {"ok": True, "minutos_extra": asistencia.minutos_extra_calculados}


@router.put("/asistencias/{asistencia_id}/revocar-autorizacion-rrhh", dependencies=[ADMIN_ACCESS])
def revocar_autorizacion_horas_extra_rrhh(asistencia_id: int, db: Session = Depends(get_db)):
    """Revoca la autorización de horas extra por parte de RRHH"""
    asistencia = db.get(RegistroAsistencia, asistencia_id)
    if not asistencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro de asistencia no encontrado")
    
    asistencia.autorizacion_horas_extra_rrhh = False
    db.commit()
    return {"ok": True}


@router.put("/asistencias/{asistencia_id}/revocar-validacion-supervisor", dependencies=[ADMIN_ACCESS])
def revocar_validacion_supervisor(asistencia_id: int, db: Session = Depends(get_db)):
    """Revoca la validación de supervisor"""
    asistencia = db.get(RegistroAsistencia, asistencia_id)
    if not asistencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro de asistencia no encontrado")
    
    asistencia.validacion_supervisor = False
    db.commit()
    return {"ok": True}


@router.put("/asistencias/{asistencia_id}/modificar-horas-extra", dependencies=[ADMIN_ACCESS])
def modificar_horas_extra(asistencia_id: int, minutos: int, db: Session = Depends(get_db)):
    """Modifica la cantidad de minutos extra calculados"""
    asistencia = db.get(RegistroAsistencia, asistencia_id)
    if not asistencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro de asistencia no encontrado")
    
    asistencia.minutos_extra_calculados = minutos
    db.commit()
    return {"ok": True, "minutos_extra": minutos}


@router.get("/bloques-horas-extra", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH, RolNombre.SUPERUSUARIO))])
def listar_bloques_horas_extra(
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    empleado_id: int | None = None,
    db: Session = Depends(get_db)
):
    """Lista los bloques de horas extra separados por tipo (ANTES_INICIO y DESPUES_FIN)"""
    from datetime import datetime

    query = select(BloqueHorasExtra).join(RegistroAsistencia)
    
    if fecha_inicio and fecha_fin:
        inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
        query = query.where(
            RegistroAsistencia.fecha_turno >= inicio,
            RegistroAsistencia.fecha_turno <= fin
        )
    
    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)
    
    bloques = db.scalars(query.order_by(BloqueHorasExtra.hora_inicio)).all()
    
    reporte = []
    for bloque in bloques:
        asistencia = bloque.asistencia
        empleado = asistencia.empleado
        reporte.append({
            "id": bloque.id,
            "asistencia_id": bloque.asistencia_id,
            "empleado_id": empleado.id,
            "nombre_empleado": empleado.nombre_completo,
            "numero_empleado": empleado.numero_empleado,
            "departamento": empleado.departamento.nombre,
            "fecha": asistencia.fecha_turno.isoformat(),
            "tipo_bloque": bloque.tipo_bloque,
            "hora_inicio": bloque.hora_inicio.isoformat(),
            "hora_fin": bloque.hora_fin.isoformat(),
            "minutos_extra": bloque.minutos_extra,
            "horas_extra": round(bloque.minutos_extra / 60, 2),
            "validado_supervisor": bloque.validacion_supervisor,
            "validado_rrhh": bloque.validacion_rrhh
        })
    
    return reporte


@router.put("/bloques-horas-extra/{bloque_id}/rechazar-rrhh", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH, RolNombre.SUPERUSUARIO))])
def rechazar_bloque_horas_extra_rrhh(bloque_id: int, db: Session = Depends(get_db)):
    """Rechaza un bloque de horas extra por parte de RRHH"""
    bloque = db.get(BloqueHorasExtra, bloque_id)
    if not bloque:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bloque de horas extra no encontrado")
    
    # Marcar como rechazado (revocar validaciones)
    bloque.validacion_supervisor = False
    bloque.validacion_rrhh = False
    db.commit()
    return {"ok": True}


# CRUD de departamentos
@router.post("/departamentos", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def crear_departamento(nombre: str, db: Session = Depends(get_db)):
    try:
        departamento = Departamento(nombre=nombre)
        db.add(departamento)
        db.commit()
        db.refresh(departamento)
        return departamento
    except Exception as e:
        db.rollback()
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El departamento ya existe")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error al crear departamento")


@router.get("/departamentos", dependencies=[ADMIN_ACCESS])
def listar_departamentos(db: Session = Depends(get_db)):
    return db.scalars(select(Departamento).order_by(Departamento.nombre)).all()


@router.get("/roles", dependencies=[ADMIN_ACCESS])
def listar_roles(db: Session = Depends(get_db)):
    roles = db.scalars(select(Rol).order_by(Rol.id)).all()
    return [
        {
            "id": r.id,
            "nombre": r.nombre
        }
        for r in roles
    ]


@router.put("/departamentos/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def actualizar_departamento(id: int, nombre: str, db: Session = Depends(get_db)):
    departamento = db.get(Departamento, id)
    if not departamento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Departamento no encontrado")
    departamento.nombre = nombre
    db.commit()
    db.refresh(departamento)
    return departamento


@router.delete("/departamentos/{id}", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR))])
def eliminar_departamento(id: int, db: Session = Depends(get_db)):
    departamento = db.get(Departamento, id)
    if not departamento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Departamento no encontrado")
    # Verificar que no tenga empleados activos
    empleados_count = db.scalar(select(Empleado).where(Empleado.departamento_id == id).with_only_columns(Empleado.id).count())
    if empleados_count > 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se puede eliminar un departamento con empleados activos")
    db.delete(departamento)
    db.commit()
    return {"ok": True}


# Endpoints para gestión de ausencias (vacaciones, incapacidades, permisos)
@router.get("/ausencias", response_model=list[AusenciaOut], dependencies=[ADMIN_ACCESS])
def obtener_ausencias(fecha: str = None, empleado_id: int = None, db: Session = Depends(get_db)):
    query = select(RegistroAusencia)
    
    if fecha:
        try:
            from datetime import date
            fecha_dt = date.fromisoformat(fecha)
            # Filtrar ausencias que incluyan la fecha especificada
            query = query.where(
                RegistroAusencia.fecha_inicio <= fecha_dt,
                RegistroAusencia.fecha_fin >= fecha_dt
            )
        except ValueError:
            pass
    
    if empleado_id:
        query = query.where(RegistroAusencia.empleado_id == empleado_id)
    
    return db.scalars(query.order_by(RegistroAusencia.fecha_inicio.desc())).all()


@router.post("/ausencias", response_model=AusenciaOut, dependencies=[ADMIN_ACCESS])
def crear_ausencia(payload: AusenciaCreate, db: Session = Depends(get_db)):
    empleado = db.get(Empleado, payload.empleado_id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    
    ausencia = RegistroAusencia(
        empleado_id=payload.empleado_id,
        tipo_ausencia=payload.tipo_ausencia,
        fecha_inicio=payload.fecha_inicio,
        fecha_fin=payload.fecha_fin,
        pagada=payload.pagada,
        porcentaje_aportacion=payload.porcentaje_aportacion,
        motivo=payload.motivo,
        fecha_registro=utc_now()
    )
    db.add(ausencia)
    db.commit()
    db.refresh(ausencia)
    return ausencia


@router.put("/ausencias/{ausencia_id}/aprobar", dependencies=[ADMIN_ACCESS])
def aprobar_ausencia(ausencia_id: int, db: Session = Depends(get_db)):
    ausencia = db.get(RegistroAusencia, ausencia_id)
    if not ausencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ausencia no encontrada")
    
    ausencia.aprobado_rrhh = True
    db.commit()
    return {"ok": True}


@router.delete("/ausencias/{ausencia_id}", dependencies=[ADMIN_ACCESS])
def eliminar_ausencia(ausencia_id: int, db: Session = Depends(get_db)):
    ausencia = db.get(RegistroAusencia, ausencia_id)
    if not ausencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ausencia no encontrada")
    db.delete(ausencia)
    db.commit()
    return {"ok": True}


# Endpoints para reportes de RRHH
@router.get("/reportes/horas-laboradas", dependencies=[ADMIN_ACCESS])
def reporte_horas_laboradas(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    from app.services import verificar_ausencia_aprobada
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        if inicio.weekday() != 4:
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        if fin.weekday() != 4:
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days=dias_adelante)
    
    query = select(RegistroAsistencia).join(Empleado).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        RegistroAsistencia.hora_entrada_real.is_not(None),
        RegistroAsistencia.hora_salida_real.is_not(None),
        Empleado.activo == True
    )
    
    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)
    
    if departamento_id:
        query = query.where(Empleado.departamento_id == departamento_id)
    
    registros = db.scalars(query.order_by(RegistroAsistencia.fecha_turno)).all()
    
    from app.services import calcular_horas_laboradas, get_empleado_turno
    from datetime import datetime
    from app.models import EstadoVisita, TipoCorreccion
    reporte = []
    for reg in registros:
        if reg.hora_entrada_real and reg.hora_salida_real:
            # Si es una visita, verificar si está autorizada por RRHH
            visita = None
            if reg.estado_registro == EstadoRegistro.VISITA_DESCANSO:
                visita = db.scalar(
                    select(Visita).where(
                        Visita.asistencia_id == reg.id,
                        Visita.estado == EstadoVisita.PAGADA,
                        Visita.fecha_autorizacion.is_not(None)
                    )
                )
            
            # Obtener horario oficial del empleado
            empleado = db.get(Empleado, reg.empleado_id)
            turno = get_empleado_turno(db, empleado, reg.fecha_turno.weekday(), reg.fecha_turno)
            
            # Calcular tiempo dentro del horario oficial
            minutos_laborados = 0
            hora_entrada_oficial = None
            hora_salida_oficial = None
            inicio_laborado = None
            fin_laborado = None
            
            if turno and turno["hora_entrada_oficial"] and turno["hora_salida_oficial"]:
                now = utc_now()
                hora_entrada_oficial = datetime.combine(reg.fecha_turno, turno["hora_entrada_oficial"], tzinfo=now.tzinfo)
                hora_salida_oficial = datetime.combine(reg.fecha_turno, turno["hora_salida_oficial"], tzinfo=now.tzinfo)
                
                hora_entrada = reg.hora_entrada_real
                hora_salida = reg.hora_salida_real
                if hora_entrada.tzinfo is None:
                    hora_entrada = hora_entrada.replace(tzinfo=now.tzinfo)
                if hora_salida.tzinfo is None:
                    hora_salida = hora_salida.replace(tzinfo=now.tzinfo)
                
                # Calcular tiempo dentro del horario oficial
                inicio_laborado = max(hora_entrada, hora_entrada_oficial)
                fin_laborado = min(hora_salida, hora_salida_oficial)
                
                if fin_laborado > inicio_laborado:
                    minutos_laborados = int((fin_laborado - inicio_laborado).total_seconds() / 60)
            
            # Si es visita autorizada, mostrarla como registro
            if visita and visita.minutos_duracion:
                minutos_registro = visita.minutos_duracion
                horas_registro = round(minutos_registro / 60, 2)
                
                # Verificar si hay ausencia aprobada en esa fecha
                ausencia = verificar_ausencia_aprobada(db, reg.empleado_id, reg.fecha_turno)
                estado_ausencia = None
                if ausencia:
                    estado_ausencia = {
                        "tipo": ausencia["tipo_ausencia"],
                        "pagada": ausencia["pagada"],
                        "porcentaje_aportacion": ausencia["porcentaje_aportacion"]
                    }
                
                reporte.append({
                    "empleado_id": reg.empleado_id,
                    "nombre_empleado": empleado.nombre_completo,
                    "numero_empleado": empleado.numero_empleado,
                    "departamento": empleado.departamento.nombre,
                    "fecha": reg.fecha_turno.isoformat(),
                    "hora_entrada": reg.hora_entrada_real.isoformat(),
                    "hora_salida": reg.hora_salida_real.isoformat(),
                    "horas_laboradas": horas_registro,
                    "estado": "VISITA_AUTORIZADA",
                    "estado_ausencia": estado_ausencia,
                    "correcciones_manuales": [],
                    "minutos_descanso": 0
                })
            
            # Si hay tiempo dentro del horario oficial, mostrarlo como registro separado
            if minutos_laborados > 0:
                # Descontar tiempo de salidas temporales que se solapen con el período laborado
                from app.models import TipoSalida
                salidas_temporales = db.scalars(
                    select(SalidaTemporal).where(
                        SalidaTemporal.asistencia_id == reg.id,
                        SalidaTemporal.tipo_salida == TipoSalida.PERMISO_PERSONAL,
                        SalidaTemporal.descuenta_tiempo == True
                    )
                ).all()
                
                minutos_descanso = 0
                for st in salidas_temporales:
                    if st.minutos_descontados:
                        minutos_descanso += st.minutos_descontados
                
                minutos_laborados -= minutos_descanso
                
                # Si después de descontar el permiso aún hay tiempo laborado, mostrarlo
                if minutos_laborados > 0:
                    horas_laboradas = round(minutos_laborados / 60, 2)
                    
                    # Verificar si hay ausencia aprobada en esa fecha
                    ausencia = verificar_ausencia_aprobada(db, reg.empleado_id, reg.fecha_turno)
                    estado_ausencia = None
                    if ausencia:
                        estado_ausencia = {
                            "tipo": ausencia["tipo_ausencia"],
                            "pagada": ausencia["pagada"],
                            "porcentaje_aportacion": ausencia["porcentaje_aportacion"]
                        }
                    
                    reporte.append({
                        "empleado_id": reg.empleado_id,
                        "nombre_empleado": empleado.nombre_completo,
                        "numero_empleado": empleado.numero_empleado,
                        "departamento": empleado.departamento.nombre,
                        "fecha": reg.fecha_turno.isoformat(),
                        "hora_entrada": inicio_laborado.isoformat(),
                        "hora_salida": fin_laborado.isoformat(),
                        "horas_laboradas": horas_laboradas,
                        "estado": "HORAS_LABORADAS",
                        "estado_ausencia": estado_ausencia,
                        "correcciones_manuales": [],
                        "minutos_descanso": minutos_descanso
                    })
    
    # Agregar correcciones manuales de tipo HORAS_LABORADAS como registros individuales
    correcciones_query = select(CorreccionManual, Empleado).join(
        Empleado, CorreccionManual.empleado_id == Empleado.id
    ).where(
        CorreccionManual.fecha >= inicio,
        CorreccionManual.fecha <= fin,
        CorreccionManual.tipo_correccion == TipoCorreccion.HORAS_LABORADAS
    )
    
    if empleado_id:
        correcciones_query = correcciones_query.where(CorreccionManual.empleado_id == empleado_id)
    
    resultados_correcciones = db.execute(correcciones_query.order_by(CorreccionManual.fecha)).all()
    
    for correccion, emp in resultados_correcciones:
        horas_correccion = round(correccion.minutos_agregados / 60, 2)
        reporte.append({
            "empleado_id": emp.id,
            "nombre_empleado": emp.nombre_completo,
            "numero_empleado": emp.numero_empleado,
            "fecha": correccion.fecha.isoformat(),
            "hora_entrada": None,
            "hora_salida": None,
            "horas_laboradas": horas_correccion,
            "estado": "CORRECCION_MANUAL",
            "estado_ausencia": None,
            "correcciones_manuales": [{"minutos": correccion.minutos_agregados, "motivo": correccion.motivo}],
            "minutos_descanso": 0
        })
    
    return reporte


@router.get("/reportes/horas-laboradas-excel", dependencies=[ADMIN_ACCESS])
def reporte_horas_laboradas_excel(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    """Reporte de horas laboradas agrupado por empleado y fecha para exportar a Excel."""
    from datetime import datetime, timedelta
    from app.services import verificar_ausencia_aprobada
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        if inicio.weekday() != 4:
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        if fin.weekday() != 4:
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days_adelante)
    
    # Obtener empleados para el reporte
    empleados_query = select(Empleado).where(Empleado.activo == True)
    if empleado_id:
        empleados_query = empleados_query.where(Empleado.id == empleado_id)
    empleados = db.scalars(empleados_query).all()
    
    from app.services import calcular_horas_laboradas
    reporte = []
    
    # Agrupar por empleado y fecha para sumar horas laboradas
    for empleado in empleados:
        fecha_actual = inicio
        while fecha_actual <= fin:
            # Obtener todos los registros del día
            registros_dia = db.scalars(
                select(RegistroAsistencia).where(
                    RegistroAsistencia.empleado_id == empleado.id,
                    RegistroAsistencia.fecha_turno == fecha_actual,
                    RegistroAsistencia.hora_entrada_real.is_not(None),
                    RegistroAsistencia.hora_salida_real.is_not(None)
                ).order_by(RegistroAsistencia.hora_entrada_real)
            ).all()
            
            # Filtrar visitas no autorizadas
            from app.models import EstadoVisita
            registros_filtrados = []
            for reg in registros_dia:
                if reg.estado_registro == "VISITA_DESCANSO":
                    visita = db.scalar(
                        select(Visita).where(
                            Visita.asistencia_id == reg.id,
                            Visita.estado == EstadoVisita.PAGADA,
                            Visita.fecha_autorizacion.is_not(None)
                        )
                    )
                    if visita:
                        # Incluir visita autorizada en la suma
                        registros_filtrados.append(reg)
                else:
                    # Incluir registros normales
                    registros_filtrados.append(reg)
            
            registros_dia = registros_filtrados
            
            # Sumar duración total de todos los registros del día
            minutos_totales = 0
            for reg in registros_dia:
                hora_entrada = reg.hora_entrada_real
                hora_salida = reg.hora_salida_real
                if hora_entrada.tzinfo is None:
                    hora_entrada = hora_entrada.replace(tzinfo=utc_now().tzinfo)
                if hora_salida.tzinfo is None:
                    hora_salida = hora_salida.replace(tzinfo=utc_now().tzinfo)
                minutos_totales += int((hora_salida - hora_entrada).total_seconds() / 60)
            
            # Obtener correcciones manuales de tipo HORAS_LABORADAS para este empleado y fecha
            from app.models import TipoCorreccion
            correcciones = db.scalars(
                select(CorreccionManual).where(
                    CorreccionManual.empleado_id == empleado.id,
                    CorreccionManual.fecha == fecha_actual,
                    CorreccionManual.tipo_correccion == TipoCorreccion.HORAS_LABORADAS
                )
            ).all()
            
            # Sumar minutos de correcciones manuales
            minutos_correcciones = sum(c.minutos_agregados for c in correcciones)
            minutos_totales += minutos_correcciones
            
            horas = minutos_totales / 60
            
            # Solo incluir si hay horas laboradas o si hay asistencia registrada
            if horas > 0:
                # Verificar si hay ausencia aprobada en esa fecha
                ausencia = verificar_ausencia_aprobada(db, empleado.id, fecha_actual)
                estado_ausencia = None
                if ausencia:
                    estado_ausencia = {
                        "tipo": ausencia["tipo_ausencia"],
                        "pagada": ausencia["pagada"],
                        "porcentaje_aportacion": ausencia["porcentaje_aportacion"]
                    }
                
                hora_entrada = None
                hora_salida = None
                if registros_dia:
                    hora_entrada = registros_dia[0].hora_entrada_real.isoformat() if registros_dia[0].hora_entrada_real else None
                    # Buscar la última salida registrada
                    for reg in reversed(registros_dia):
                        if reg.hora_salida_real:
                            # Verificar si hay bloques de horas extra DESPUES_FIN para usar la hora de corte
                            bloques_extra = db.scalars(
                                select(BloqueHorasExtra).where(
                                    BloqueHorasExtra.asistencia_id == reg.id,
                                    BloqueHorasExtra.tipo_bloque == 'DESPUES_FIN'
                                )
                            ).all()
                            if bloques_extra:
                                hora_salida = bloques_extra[0].hora_inicio.isoformat()
                            else:
                                hora_salida = reg.hora_salida_real.isoformat()
                            break
                
                reporte.append({
                    "empleado_id": empleado.id,
                    "nombre_empleado": empleado.nombre_completo,
                    "numero_empleado": empleado.numero_empleado,
                    "fecha": fecha_actual.isoformat(),
                    "hora_entrada": hora_entrada,
                    "hora_salida": hora_salida,
                    "horas_laboradas": round(horas, 2),
                    "estado_ausencia": estado_ausencia,
                    "correcciones_manuales": [{"minutos": c.minutos_agregados, "motivo": c.motivo} for c in correcciones],
                    "minutos_descanso": 0
                })
            
            fecha_actual += timedelta(days=1)
    
    return reporte


@router.get("/reportes/horas-extra", dependencies=[ADMIN_ACCESS])
def reporte_horas_extra(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    from app.models import Visita, EstadoVisita
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        if inicio.weekday() != 4:
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        if fin.weekday() != 4:
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days=dias_adelante)
    
    reporte = []
    
    # 1. Obtener bloques de horas extra de la base de datos (detectados por calcular_y_registrar_bloques_horas_extra)
    query_bloques = select(BloqueHorasExtra, RegistroAsistencia).join(
        RegistroAsistencia, BloqueHorasExtra.asistencia_id == RegistroAsistencia.id
    ).join(
        Empleado, RegistroAsistencia.empleado_id == Empleado.id
    ).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        Empleado.activo == True
    )
    
    if empleado_id:
        query_bloques = query_bloques.where(RegistroAsistencia.empleado_id == empleado_id)
    
    if departamento_id:
        query_bloques = query_bloques.where(Empleado.departamento_id == departamento_id)
    
    resultados_bloques = db.execute(query_bloques.order_by(RegistroAsistencia.fecha_turno, BloqueHorasExtra.hora_inicio)).all()
    
    for bloque, reg in resultados_bloques:
        empleado = reg.empleado
        # Determinar estado basado en validaciones
        if bloque.validacion_rrhh:
            estado = "Autorizado RRHH"
        elif bloque.validacion_supervisor:
            estado = "Validado Supervisor"
        else:
            estado = "Pendiente"
        
        reporte.append({
            "tipo": bloque.tipo_bloque,
            "empleado_id": empleado.id,
            "nombre_empleado": empleado.nombre_completo,
            "numero_empleado": empleado.numero_empleado,
            "departamento": empleado.departamento.nombre,
            "fecha": reg.fecha_turno.isoformat(),
            "bloque_id": bloque.id,
            "hora_inicio": bloque.hora_inicio.isoformat() if bloque.hora_inicio else None,
            "hora_fin": bloque.hora_fin.isoformat() if bloque.hora_fin else None,
            "minutos_extra": bloque.minutos_extra,
            "horas_extra": round(bloque.minutos_extra / 60, 2),
            "estado": estado,
            "validado_supervisor": bloque.validacion_supervisor,
            "validado_rrhh": bloque.validacion_rrhh
        })
    
    # 2. Obtener visitas pagadas autorizadas por RRHH
    query_visitas = select(Visita, RegistroAsistencia).join(
        RegistroAsistencia, Visita.asistencia_id == RegistroAsistencia.id
    ).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        Visita.estado == EstadoVisita.PAGADA,
        Visita.autorizado_por.is_not(None)
    )
    
    if empleado_id:
        query_visitas = query_visitas.where(RegistroAsistencia.empleado_id == empleado_id)
    
    resultados_visitas = db.execute(query_visitas.order_by(RegistroAsistencia.fecha_turno, Visita.fecha_visita)).all()
    
    for visita, reg in resultados_visitas:
        empleado = reg.empleado
        if departamento_id and empleado.departamento_id != departamento_id:
            continue
        if not empleado.activo:
            continue
        minutos_visita = visita.minutos_duracion if visita.minutos_duracion else 0
        reporte.append({
            "tipo": "Visita_Pagada",
            "empleado_id": empleado.id,
            "nombre_empleado": empleado.nombre_completo,
            "numero_empleado": empleado.numero_empleado,
            "departamento": empleado.departamento.nombre,
            "fecha": reg.fecha_turno.isoformat(),
            "visita_id": visita.id,
            "fecha_visita": visita.fecha_visita.isoformat(),
            "hora_inicio": visita.hora_inicio.isoformat() if visita.hora_inicio else None,
            "hora_fin": visita.hora_fin.isoformat() if visita.hora_fin else None,
            "minutos_extra": minutos_visita,
            "horas_extra": round(minutos_visita / 60, 2),
            "motivo": visita.motivo,
            "autorizado_por": visita.autorizado_por,
            "fecha_autorizacion": visita.fecha_autorizacion.isoformat() if visita.fecha_autorizacion else None
        })
    
    # 3. Obtener correcciones manuales de horas extra
    from app.models import CorreccionManual, TipoCorreccion
    query_correcciones = select(CorreccionManual, Empleado).join(
        Empleado, CorreccionManual.empleado_id == Empleado.id
    ).where(
        CorreccionManual.fecha >= inicio,
        CorreccionManual.fecha <= fin,
        CorreccionManual.tipo_correccion == TipoCorreccion.HORAS_EXTRA
    )
    
    if empleado_id:
        query_correcciones = query_correcciones.where(CorreccionManual.empleado_id == empleado_id)
    
    resultados_correcciones = db.execute(query_correcciones.order_by(CorreccionManual.fecha)).all()
    
    for correccion, emp in resultados_correcciones:
        reporte.append({
            "tipo": "Entrada_Manual",
            "empleado_id": emp.id,
            "fecha": correccion.fecha.isoformat(),
            "correccion_id": correccion.id,
            "minutos_extra": correccion.minutos_agregados,
            "horas_extra": round(correccion.minutos_agregados / 60, 2),
            "motivo": correccion.motivo
        })
    
    # Ordenar reporte por fecha y tipo
    reporte.sort(key=lambda x: (x["fecha"], x["tipo"]))
    
    return reporte


@router.get("/reportes/asistencias", dependencies=[ADMIN_ACCESS])
def reporte_asistencias(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    from app.services import verificar_ausencia_aprobada
    from app.models import TipoAusencia
    
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        # Inicio: viernes 8am
        if inicio.weekday() != 4:  # 4 = viernes
            # Si no es viernes, buscar el viernes anterior
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        inicio_dt = datetime.combine(inicio, datetime.min.time(), tzinfo=utc_now().tzinfo) + timedelta(hours=8)
        
        # Fin: siguiente viernes 8am
        if fin.weekday() != 4:  # 4 = viernes
            # Si no es viernes, buscar el siguiente viernes
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days=dias_adelante)
        fin_dt = datetime.combine(fin, datetime.min.time(), tzinfo=utc_now().tzinfo) + timedelta(hours=8)
        
        # Ajustar a fechas para el reporte (usar las fechas completas)
        inicio = inicio_dt.date()
        fin = fin_dt.date()
    
    # Filtrar empleados según parámetro - incluir TODOS los empleados activos
    empleados_query = select(Empleado).where(Empleado.activo == True)
    if empleado_id:
        empleados_query = empleados_query.where(Empleado.id == empleado_id)
    if departamento_id:
        empleados_query = empleados_query.where(Empleado.departamento_id == departamento_id)
    
    empleados = db.scalars(empleados_query).all()
    
    reporte = []
    
    # Generar reporte día por día para cada empleado
    for empleado in empleados:
        fecha_actual = inicio
        while fecha_actual <= fin:
            # Verificar si hay registro de asistencia
            asistencia = db.scalar(
                select(RegistroAsistencia).where(
                    RegistroAsistencia.empleado_id == empleado.id,
                    RegistroAsistencia.fecha_turno == fecha_actual
                )
            )
            
            if asistencia and asistencia.hora_entrada_real:
                # Hay registro de asistencia
                leyenda = None
                if asistencia.estado_registro:
                    leyenda = asistencia.estado_registro.value
                
                # Calcular horas laboradas
                from app.services import calcular_horas_laboradas
                horas_info = calcular_horas_laboradas(db, empleado.id, fecha_actual)
                horas_laboradas = round(horas_info.get("minutos_laborados", 0) / 60, 2)
                horas_extra = round(horas_info.get("minutos_extra", 0) / 60, 2)
                # Filtrar correcciones de tipo Horas_Laboradas y Permiso
                correcciones = [c for c in horas_info.get("correcciones", []) if c["tipo"] in ["Horas_Laboradas", "Permiso"]]
                minutos_descanso = horas_info.get("minutos_descanso", 0)
                
                reporte.append({
                    "empleado_id": empleado.id,
                    "nombre_empleado": empleado.nombre_completo,
                    "numero_empleado": empleado.numero_empleado,
                    "fecha": fecha_actual.isoformat(),
                    "hora_entrada": asistencia.hora_entrada_real.isoformat() if asistencia.hora_entrada_real else None,
                    "hora_salida": asistencia.hora_salida_real.isoformat() if asistencia.hora_salida_real else None,
                    "estado_registro": leyenda,
                    "leyenda": leyenda,
                    "horas_laboradas": horas_laboradas,
                    "horas_extra": horas_extra,
                    "correcciones_manuales": correcciones,
                    "minutos_descanso": minutos_descanso
                })
            else:
                # No hay registro de asistencia, verificar si hay ausencia aprobada
                ausencia = verificar_ausencia_aprobada(db, empleado.id, fecha_actual)
                if ausencia:
                    # Hay ausencia aprobada
                    if ausencia["tipo_ausencia"] == TipoAusencia.VACACIONES.value:
                        leyenda = "V"  # Vacaciones
                    elif ausencia["tipo_ausencia"] == TipoAusencia.INCAPACIDAD.value:
                        leyenda = "I"  # Incapacidad
                    else:
                        leyenda = "A"  # Otro tipo de ausencia
                    
                    reporte.append({
                        "empleado_id": empleado.id,
                        "nombre_empleado": empleado.nombre_completo,
                        "numero_empleado": empleado.numero_empleado,
                        "fecha": fecha_actual.isoformat(),
                        "hora_entrada": None,
                        "hora_salida": None,
                        "estado_registro": None,
                        "leyenda": leyenda,
                        "ausencia": ausencia,
                        "horas_laboradas": 0,
                        "horas_extra": 0
                    })
                else:
                    # No hay ausencia ni asistencia
                    # Verificar si es día de descanso del empleado
                    from app.services import get_empleado_turno
                    turno_dia = get_empleado_turno(db, empleado, fecha_actual.weekday(), fecha_actual)
                    es_dia_descanso = turno_dia is not None and turno_dia.get("es_descanso", False)

                    reporte.append({
                        "empleado_id": empleado.id,
                        "nombre_empleado": empleado.nombre_completo,
                        "numero_empleado": empleado.numero_empleado,
                        "fecha": fecha_actual.isoformat(),
                        "hora_entrada": None,
                        "hora_salida": None,
                        "estado_registro": None,
                        "leyenda": "D" if es_dia_descanso else "F",  # D=Descanso, F=Falta injustificada
                        "horas_laboradas": 0,
                        "horas_extra": 0
                    })
            
            fecha_actual += timedelta(days=1)
    
    return reporte


@router.post("/procesar-visitas-vencidas", dependencies=[ADMIN_ACCESS])
def procesar_visitas_vencidas(db: Session = Depends(get_db)):
    """Procesa visitas vencidas (pendientes por más de 2 días) y las marca como NO_PAGADA."""
    from app.services import procesar_visitas_vencidas
    
    cantidad = procesar_visitas_vencidas(db)
    return {"ok": True, "visitas_procesadas": cantidad}


@router.get("/reportes/horas-laboradas/excel", dependencies=[ADMIN_ACCESS])
def exportar_horas_laboradas_excel(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        if inicio.weekday() != 4:
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        if fin.weekday() != 4:
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days=dias_adelante)

    query = select(RegistroAsistencia).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        RegistroAsistencia.hora_entrada_real.is_not(None),
        RegistroAsistencia.hora_salida_real.is_not(None)
    )

    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)

    registros = db.scalars(query.order_by(RegistroAsistencia.fecha_turno)).all()

    # Obtener todos los departamentos o solo el filtrado
    departamentos_query = select(Departamento)
    if departamento_id:
        departamentos_query = departamentos_query.where(Departamento.id == departamento_id)
    departamentos = db.scalars(departamentos_query).all()

    # Obtener todos los empleados activos
    empleados_query = select(Empleado).where(Empleado.activo == True)
    if empleado_id:
        empleados_query = empleados_query.where(Empleado.id == empleado_id)
    if departamento_id:
        empleados_query = empleados_query.where(Empleado.departamento_id == departamento_id)
    empleados = db.scalars(empleados_query).all()

    # Calcular todos los días en el rango
    dias_en_rango = []
    fecha_actual = inicio
    while fecha_actual <= fin:
        dias_en_rango.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    # Nombres de días en español (sin acentos para compatibilidad Excel)
    dias_semana = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']

    wb = Workbook()
    wb.remove(wb.active)

    # Crear hoja por departamento
    for dept in departamentos:
        dept_empleados = [e for e in empleados if e.departamento_id == dept.id]
        if not dept_empleados:
            continue

        # Limpiar nombre de hoja: remover caracteres especiales y limitar longitud
        nombre_hoja = dept.nombre[:31].replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
        ws = wb.create_sheet(title=nombre_hoja)

        # Header con días
        headers = ['Número', 'Colaborador'] + [f"{d.strftime('%d/%m')} ({dias_semana[d.weekday()]})" for d in dias_en_rango] + ['TOTAL', 'Minutos Permiso', 'Correcciones']
        ws.append(headers)

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data por empleado
        for emp in dept_empleados:
            fila = [emp.numero_empleado, emp.nombre_completo]
            total_horas = 0
            total_minutos_permiso = 0
            correcciones_texto = []

            for fecha in dias_en_rango:
                # Usar calcular_horas_laboradas para TODOS los días:
                # incluye horas trabajadas, correcciones manuales y ausencias
                # (vacaciones/incapacidades/permisos pagados cuentan como horas)
                from app.services import calcular_horas_laboradas
                horas_info = calcular_horas_laboradas(db, emp.id, fecha)
                horas = round(horas_info.get("minutos_laborados", 0) / 60, 2)
                minutos_descanso = horas_info.get("minutos_descanso", 0)
                total_minutos_permiso += minutos_descanso
                if horas > 0:
                    fila.append(horas)
                    total_horas += horas
                else:
                    fila.append(0)
                
                # Recopilar correcciones manuales para mostrar en columna (Horas_Laboradas y Permiso)
                correcciones = [c for c in horas_info.get("correcciones", []) if c["tipo"] in ["Horas_Laboradas", "Permiso"]]
                for c in correcciones:
                    if c['tipo'] == 'Permiso':
                        correcciones_texto.append(f"{fecha.strftime('%d/%m')}: {c['tipo']} {c['minutos']:.1f}min")
                    else:
                        correcciones_texto.append(f"{fecha.strftime('%d/%m')}: {c['tipo']} {c['minutos']:+.1f}min")

            fila.append(round(total_horas, 2))
            fila.append(total_minutos_permiso)
            fila.append('; '.join(correcciones_texto) if correcciones_texto else '')
            ws.append(fila)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Si no hay datos, crear hoja vacía
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Sin Datos")
        ws.append(['No hay registros para el período seleccionado'])

    filename = f"horas_laboradas_{fecha_inicio}_{fecha_fin}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/reportes/horas-extra/excel", dependencies=[ADMIN_ACCESS])
def exportar_horas_extra_excel(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    corte_semanal: bool = False,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    
    # Si es corte semanal, ajustar las fechas a viernes 8am
    if corte_semanal:
        if inicio.weekday() != 4:
            dias_atras = (inicio.weekday() - 4) % 7
            inicio = inicio - timedelta(days=dias_atras)
        if fin.weekday() != 4:
            dias_adelante = (4 - fin.weekday()) % 7
            fin = fin + timedelta(days=dias_adelante)

    query = select(RegistroAsistencia).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        RegistroAsistencia.minutos_extra_calculados > 0,
        RegistroAsistencia.autorizacion_horas_extra_rrhh == True
    )

    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)

    registros = db.scalars(query.order_by(RegistroAsistencia.fecha_turno)).all()

    departamentos_query = select(Departamento)
    if departamento_id:
        departamentos_query = departamentos_query.where(Departamento.id == departamento_id)
    departamentos = db.scalars(departamentos_query).all()

    empleados_query = select(Empleado).where(Empleado.activo == True)
    if empleado_id:
        empleados_query = empleados_query.where(Empleado.id == empleado_id)
    if departamento_id:
        empleados_query = empleados_query.where(Empleado.departamento_id == departamento_id)
    empleados = db.scalars(empleados_query).all()

    dias_en_rango = []
    fecha_actual = inicio
    while fecha_actual <= fin:
        dias_en_rango.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    dias_semana = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']

    wb = Workbook()
    wb.remove(wb.active)

    for dept in departamentos:
        dept_empleados = [e for e in empleados if e.departamento_id == dept.id]
        if not dept_empleados:
            continue

        # Limpiar nombre de hoja: remover caracteres especiales y limitar longitud
        nombre_hoja = dept.nombre[:31].replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
        ws = wb.create_sheet(title=nombre_hoja)

        headers = ['Número', 'Colaborador'] + [f"{d.strftime('%d/%m')} ({dias_semana[d.weekday()]})" for d in dias_en_rango] + ['TOTAL', 'Entradas Manuales']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for emp in dept_empleados:
            fila = [emp.numero_empleado, emp.nombre_completo]
            total_horas = 0
            entradas_manuales_texto = []

            for fecha in dias_en_rango:
                # Calcular horas extra del día usando bloques_extra detectados + autorizados + visitas + correcciones
                minutos_extra = 0
                
                # 1. Bloques de horas extra detectados por calcular_horas_laboradas
                from app.services import calcular_horas_laboradas
                horas_info = calcular_horas_laboradas(db, emp.id, fecha)
                for bloque in horas_info.get("bloques_extra", []):
                    minutos_extra += bloque["minutos"]
                
                # 2. Bloques de visitas de ausencias (requieren validación RRHH)
                for bloque in horas_info.get("bloques_visitas", []):
                    minutos_extra += bloque["minutos"]
                
                # 3. Bloques de horas extra autorizados por RRHH
                asistencia = db.scalar(
                    select(RegistroAsistencia).where(
                        RegistroAsistencia.empleado_id == emp.id,
                        RegistroAsistencia.fecha_turno == fecha
                    )
                )
                if asistencia:
                    bloques = db.scalars(
                        select(BloqueHorasExtra).where(
                            BloqueHorasExtra.asistencia_id == asistencia.id,
                            BloqueHorasExtra.validacion_rrhh == True  # Solo requiere validación de RRHH
                        )
                    ).all()
                    for bloque in bloques:
                        minutos_extra += bloque.minutos_extra
                
                # 4. Visitas pagadas autorizadas por RRHH
                visitas = db.scalars(
                    select(Visita).join(
                        RegistroAsistencia, Visita.asistencia_id == RegistroAsistencia.id
                    ).where(
                        RegistroAsistencia.empleado_id == emp.id,
                        RegistroAsistencia.fecha_turno == fecha,
                        Visita.estado == EstadoVisita.PAGADA,
                        Visita.autorizado_por.is_not(None)
                    )
                ).all()
                for visita in visitas:
                    minutos_extra += visita.minutos_duracion if visita.minutos_duracion else 0
                
                # 5. Correcciones manuales de horas extra (automáticamente autorizadas)
                correcciones = db.scalars(
                    select(CorreccionManual).where(
                        CorreccionManual.empleado_id == emp.id,
                        CorreccionManual.fecha == fecha,
                        CorreccionManual.tipo_correccion == TipoCorreccion.HORAS_EXTRA
                    )
                ).all()
                for correccion in correcciones:
                    minutos_extra += correccion.minutos_agregados
                    # Agregar a la lista de entradas manuales
                    entradas_manuales_texto.append(f"{fecha.strftime('%d/%m')}: {correccion.minutos_agregados:+.1f}min - {correccion.motivo}")
                
                if minutos_extra > 0:
                    horas = minutos_extra / 60
                    fila.append(round(horas, 2))
                    total_horas += horas
                else:
                    fila.append(0)

            fila.append(round(total_horas, 2))
            fila.append('; '.join(entradas_manuales_texto) if entradas_manuales_texto else '')
            ws.append(fila)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Sin Datos")
        ws.append(['No hay registros de horas extra para el período seleccionado'])

    filename = f"horas_extra_{fecha_inicio}_{fecha_fin}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/reportes/asistencias/excel", dependencies=[ADMIN_ACCESS])
def exportar_asistencias_excel(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    db: Session = Depends(get_db)
):
    from datetime import datetime, timedelta
    from app.services import verificar_ausencia_aprobada
    from app.models import TipoAusencia
    inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

    # Incluir TODOS los registros de asistencia
    query = select(RegistroAsistencia).where(
        RegistroAsistencia.fecha_turno >= inicio,
        RegistroAsistencia.fecha_turno <= fin,
        RegistroAsistencia.hora_entrada_real.is_not(None)
    )

    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)

    registros = db.scalars(query.order_by(RegistroAsistencia.fecha_turno)).all()

    departamentos_query = select(Departamento)
    if departamento_id:
        departamentos_query = departamentos_query.where(Departamento.id == departamento_id)
    departamentos = db.scalars(departamentos_query).all()

    # Incluir TODOS los empleados activos
    empleados_query = select(Empleado).where(Empleado.activo == True)
    if empleado_id:
        empleados_query = empleados_query.where(Empleado.id == empleado_id)
    if departamento_id:
        empleados_query = empleados_query.where(Empleado.departamento_id == departamento_id)
    empleados = db.scalars(empleados_query).all()

    dias_en_rango = []
    fecha_actual = inicio
    while fecha_actual <= fin:
        dias_en_rango.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    dias_semana = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']

    wb = Workbook()
    wb.remove(wb.active)

    for dept in departamentos:
        dept_empleados = [e for e in empleados if e.departamento_id == dept.id]
        if not dept_empleados:
            continue

        # Limpiar nombre de hoja: remover caracteres especiales y limitar longitud
        nombre_hoja = dept.nombre[:31].replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
        ws = wb.create_sheet(title=nombre_hoja)

        headers = ['Número', 'Colaborador'] + [f"{d.strftime('%d/%m')} ({dias_semana[d.weekday()]})" for d in dias_en_rango] + ['TOTAL']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for emp in dept_empleados:
            fila = [emp.numero_empleado, emp.nombre_completo]
            total_asistencias = 0

            for fecha in dias_en_rango:
                reg = next((r for r in registros if r.empleado_id == emp.id and r.fecha_turno == fecha), None)
                if reg and reg.hora_entrada_real:
                    fila.append('SI')
                    total_asistencias += 1
                else:
                    # Verificar si hay ausencia aprobada
                    ausencia = verificar_ausencia_aprobada(db, emp.id, fecha)
                    if ausencia:
                        if ausencia["tipo_ausencia"] == TipoAusencia.VACACIONES.value:
                            fila.append('V')  # Vacaciones
                        elif ausencia["tipo_ausencia"] == TipoAusencia.INCAPACIDAD.value:
                            fila.append('I')  # Incapacidad
                        else:
                            fila.append('A')  # Otro tipo de ausencia
                    else:
                        # Verificar si es día de descanso del empleado
                        from app.services import get_empleado_turno
                        turno_dia = get_empleado_turno(db, emp, fecha.weekday(), fecha)
                        es_dia_descanso = turno_dia is not None and turno_dia.get("es_descanso", False)
                        fila.append('D' if es_dia_descanso else 'NO')  # D=Descanso, NO=Falta

            fila.append(total_asistencias)
            ws.append(fila)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="Sin Datos")
        ws.append(['No hay registros de asistencias para el período seleccionado'])

    filename = f"asistencias_{fecha_inicio}_{fecha_fin}.xlsx"
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.put("/bloques-horas-extra/{bloque_id}/aprobar-supervisor", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.SUPERVISOR))])
def aprobar_bloque_horas_extra_supervisor(bloque_id: int, db: Session = Depends(get_db), current_user: UsuarioSistema = Depends(get_current_user)):
    """Aprueba un bloque de horas extra a nivel de supervisor"""
    bloque = db.get(BloqueHorasExtra, bloque_id)
    if not bloque:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bloque de horas extra no encontrado")
    
    bloque.validacion_supervisor = True
    db.commit()
    return {"message": "Bloque de horas extra aprobado por supervisor"}


@router.put("/bloques-horas-extra/{bloque_id}/aprobar-rrhh", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def aprobar_bloque_horas_extra_rrhh(bloque_id: int, db: Session = Depends(get_db)):
    """Aprueba un bloque de horas extra a nivel de RRHH.
    RRHH puede autorizar bloques aunque el supervisor no los haya validado.
    Solo los bloques autorizados por RRHH se incluyen en reportes y Excel."""
    bloque = db.get(BloqueHorasExtra, bloque_id)
    if not bloque:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bloque de horas extra no encontrado")
    
    bloque.validacion_rrhh = True
    db.commit()
    return {"message": "Bloque de horas extra aprobado por RRHH"}


@router.put("/bloques-horas-extra/{bloque_id}/actualizar-minutos", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH, RolNombre.SUPERVISOR))])
def actualizar_minutos_bloque_horas_extra(bloque_id: int, minutos_extra: int, db: Session = Depends(get_db)):
    """Actualiza los minutos extra de un bloque de horas extra.
    Tanto supervisor como RRHH pueden modificar la cantidad de tiempo extra autorizado."""
    bloque = db.get(BloqueHorasExtra, bloque_id)
    if not bloque:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bloque de horas extra no encontrado")
    
    if minutos_extra < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Los minutos extra no pueden ser negativos")
    
    bloque.minutos_extra = minutos_extra
    db.commit()
    db.refresh(bloque)
    return {"message": "Minutos extra actualizados", "minutos_extra": bloque.minutos_extra}


@router.put("/incidencias/horas-extra/{asistencia_id}/aprobar-rrhh", dependencies=[Depends(require_roles(RolNombre.ADMINISTRADOR, RolNombre.RRHH))])
def aprobar_horas_extra_rrhh(asistencia_id: int, db: Session = Depends(get_db)):
    """Aprueba horas extra y bloquea el registro para evitar modificaciones"""
    from app.core.time import utc_now
    asistencia = db.get(RegistroAsistencia, asistencia_id)
    if not asistencia:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asistencia no encontrada")
    if not asistencia.validacion_supervisor:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El registro debe ser validado por supervisor primero")
    asistencia.validacion_rrhh = True
    db.commit()
    db.refresh(asistencia)
    return asistencia


@router.get("/reportes/salidas-temporales", dependencies=[ADMIN_ACCESS])
def reporte_salidas_temporales(
    fecha_inicio: str,
    fecha_fin: str,
    empleado_id: int | None = None,
    departamento_id: int | None = None,
    db: Session = Depends(get_db),
):
    """Reporte detallado de salidas temporales (comer, mandado, permisos)"""
    from datetime import datetime, timedelta
    from app.core.time import utc_now

    fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

    query = (
        select(SalidaTemporal, RegistroAsistencia, Empleado)
        .join(RegistroAsistencia, SalidaTemporal.asistencia_id == RegistroAsistencia.id)
        .join(Empleado, RegistroAsistencia.empleado_id == Empleado.id)
        .where(RegistroAsistencia.fecha_turno >= fecha_inicio_dt)
        .where(RegistroAsistencia.fecha_turno <= fecha_fin_dt)
        .where(Empleado.activo == True)
    )

    if empleado_id:
        query = query.where(RegistroAsistencia.empleado_id == empleado_id)

    if departamento_id:
        query = query.where(Empleado.departamento_id == departamento_id)

    resultados = db.execute(query.order_by(SalidaTemporal.hora_salida.desc())).all()

    reporte = []
    for salida, asistencia, empleado in resultados:
        duracion_minutos = None
        if salida.hora_regreso:
            duracion_minutos = int((salida.hora_regreso - salida.hora_salida).total_seconds() / 60)

        reporte.append({
            "id": salida.id,
            "empleado_id": empleado.id,
            "nombre_empleado": empleado.nombre_completo,
            "numero_empleado": empleado.numero_empleado,
            "puesto": empleado.puesto,
            "tipo_salida": salida.tipo_salida.value,
            "hora_salida": salida.hora_salida,
            "hora_regreso": salida.hora_regreso,
            "estado_salida": salida.estado_salida.value,
            "duracion_minutos": duracion_minutos,
            "minutos_descontados": salida.minutos_descontados,
            "descuenta_tiempo": salida.descuenta_tiempo,
            "fecha_turno": asistencia.fecha_turno
        })

    return reporte


# ============ CORRECCIONES MANUALES ============

@router.get("/correcciones-manuales", response_model=list[CorreccionManualOut], dependencies=[ADMIN_ACCESS])
def obtener_correcciones(fecha_inicio: str | None = None, fecha_fin: str | None = None, fecha: str | None = None, empleado_id: int | None = None, db: Session = Depends(get_db)):
    """Obtiene todas las correcciones manuales en un rango de fechas y opcionalmente por empleado"""
    from datetime import date
    
    query = select(CorreccionManual)
    
    # Filtro por fecha exacta (fecha asignada)
    if fecha:
        try:
            fecha_dt = date.fromisoformat(fecha)
            query = query.where(CorreccionManual.fecha == fecha_dt)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato de fecha inválido")
    
    # Filtro por rango de fechas
    if fecha_inicio:
        try:
            fecha_inicio_dt = date.fromisoformat(fecha_inicio)
            query = query.where(CorreccionManual.fecha >= fecha_inicio_dt)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato de fecha_inicio inválido")
    
    if fecha_fin:
        try:
            fecha_fin_dt = date.fromisoformat(fecha_fin)
            query = query.where(CorreccionManual.fecha <= fecha_fin_dt)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Formato de fecha_fin inválido")
    
    if empleado_id:
        query = query.where(CorreccionManual.empleado_id == empleado_id)
    
    correcciones = db.scalars(query.order_by(CorreccionManual.fecha_registro.desc())).all()
    return correcciones


@router.post("/correcciones-manuales", response_model=CorreccionManualOut, dependencies=[ADMIN_ACCESS])
def crear_correccion(payload: CorreccionManualCreate, db: Session = Depends(get_db), current_user: UsuarioSistema = Depends(get_current_user)):
    """Crea una corrección manual para un empleado"""
    empleado = db.get(Empleado, payload.empleado_id)
    if not empleado:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Colaborador no encontrado")
    
    correccion = CorreccionManual(
        empleado_id=payload.empleado_id,
        fecha=payload.fecha,
        tipo_correccion=payload.tipo_correccion,
        minutos_agregados=payload.minutos_agregados,
        motivo=payload.motivo,
        autorizado_por=current_user.id,
        fecha_registro=utc_now()
    )
    db.add(correccion)
    db.commit()
    db.refresh(correccion)
    return correccion


@router.delete("/correcciones-manuales/{correccion_id}", dependencies=[ADMIN_ACCESS])
def eliminar_correccion(correccion_id: int, db: Session = Depends(get_db)):
    """Elimina una corrección manual"""
    correccion = db.get(CorreccionManual, correccion_id)
    if not correccion:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Corrección no encontrada")
    
    db.delete(correccion)
    db.commit()
    return {"message": "Corrección eliminada"}
