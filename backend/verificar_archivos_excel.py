"""
Verificar contenido de archivos Excel generados
"""
from openpyxl import load_workbook
import os

output_dir = os.path.join(os.path.dirname(__file__), "reportes_excel_prueba")

archivos = [
    "horas_laboradas_2026-06-08_2026-06-15.xlsx",
    "horas_extra_2026-06-08_2026-06-15.xlsx",
    "asistencias_2026-06-08_2026-06-15.xlsx"
]

for archivo in archivos:
    filepath = os.path.join(output_dir, archivo)
    print(f"\n{'='*60}")
    print(f"Archivo: {archivo}")
    print(f"{'='*60}")
    
    try:
        wb = load_workbook(filepath)
        print(f"Hojas: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Hoja: {sheet_name} ---")
            print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
            
            # Mostrar primeras 5 filas
            print("\nPrimeras filas:")
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if i <= 5:
                    print(f"  Fila {i}: {row}")
                else:
                    break
            
            if ws.max_row > 5:
                print(f"  ... ({ws.max_row - 5} filas mas)")
        
        print(f"\nOK: Archivo valido y legible")
        
    except Exception as e:
        print(f"ERROR: {e}")

print(f"\n{'='*60}")
print("Verificacion completada")
print(f"{'='*60}")
