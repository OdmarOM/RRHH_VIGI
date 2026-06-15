"""
Verificar especificamente el archivo de asistencias
"""
from openpyxl import load_workbook
import os

output_dir = os.path.join(os.path.dirname(__file__), "reportes_excel_prueba")
filepath = os.path.join(output_dir, "asistencias_2026-06-08_2026-06-15.xlsx")

print(f"Verificando: {filepath}")

try:
    wb = load_workbook(filepath)
    print(f"Hojas: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Hoja: {sheet_name} ---")
        print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
        
        # Mostrar todas las filas
        print("\nTodas las filas:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            print(f"  Fila {i}: {row}")
    
    print(f"\nOK: Archivo valido y legible")
    
except Exception as e:
    print(f"ERROR: {e}")
